from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADERS = [
    "ID",
    "Fecha",
    "Tipo",
    "Cantidad",
    "Moneda",
    "Categoria",
    "Subcategoria",
    "Concepto",
    "Metodo de pago",
    "Cuenta",
    "Nota",
]


def normalize_cell_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    return value


def movement_row(row: dict[str, Any]) -> list[Any]:
    return [
        row.get("id"),
        normalize_cell_value(row.get("created_at")),
        row.get("tipo"),
        normalize_cell_value(row.get("cantidad")),
        row.get("moneda"),
        row.get("categoria"),
        row.get("subcategoria"),
        row.get("concepto"),
        row.get("metodo_pago"),
        row.get("cuenta"),
        row.get("nota"),
    ]


def movement_totals(rows: list[dict[str, Any]]) -> dict[str, float]:
    expenses = 0.0
    income = 0.0

    for row in rows:
        amount = float(row.get("cantidad") or 0)
        movement_type = str(row.get("tipo") or "").lower()
        if movement_type in {"gasto", "gastos", "expense", "egreso", "egresos"}:
            expenses += amount
        elif movement_type in {"ingreso", "ingresos", "income"}:
            income += amount

    return {
        "expenses": expenses,
        "income": income,
        "balance": income - expenses,
    }


def build_export_workbook(
    rows: list[dict[str, Any]],
    filters: dict[str, str | None],
) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    movements = workbook.create_sheet("Movimientos")

    create_summary_sheet(summary, rows, filters)
    create_movements_sheet(movements, rows)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def create_summary_sheet(sheet, rows: list[dict[str, Any]], filters: dict[str, str | None]) -> None:
    totals = movement_totals(rows)
    generated_at = datetime.now().replace(microsecond=0)

    sheet["A1"] = "Exportacion de movimientos"
    sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="111827")
    sheet.merge_cells("A1:D1")

    summary_rows = [
        ("Generado", generated_at),
        ("Movimientos exportados", len(rows)),
        ("Gastos", totals["expenses"]),
        ("Ingresos", totals["income"]),
        ("Balance", totals["balance"]),
        ("Desde", filters.get("start_date") or "Sin filtro"),
        ("Hasta", filters.get("end_date") or "Sin filtro"),
        ("Categoria", filters.get("category") or "Todas"),
        ("Tipo", filters.get("movement_type") or "Todos"),
    ]

    for index, (label, value) in enumerate(summary_rows, start=3):
        sheet.cell(row=index, column=1, value=label)
        sheet.cell(row=index, column=2, value=value)
        sheet.cell(row=index, column=1).font = Font(bold=True)

    for row in range(5, 8):
        sheet.cell(row=row, column=2).number_format = '#,##0.00 "EUR"'

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 28


def create_movements_sheet(sheet, rows: list[dict[str, Any]]) -> None:
    sheet.append(HEADERS)

    for row in rows:
        sheet.append(movement_row(row))

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for cell in sheet["B"][1:]:
        if isinstance(cell.value, (date, datetime)):
            cell.number_format = "yyyy-mm-dd hh:mm"

    for cell in sheet["D"][1:]:
        cell.number_format = "#,##0.00"

    widths = {
        "A": 10,
        "B": 20,
        "C": 14,
        "D": 14,
        "E": 10,
        "F": 20,
        "G": 20,
        "H": 34,
        "I": 18,
        "J": 18,
        "K": 40,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    if rows:
        table = Table(displayName="MovimientosTable", ref=sheet.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        sheet.add_table(table)

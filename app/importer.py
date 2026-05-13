import csv
import io
import json
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

from app.database import execute_returning_id
from app.models import ImportCommitRequest, MovementPayload
from app.movements import create_movement


FIELD_ALIASES = {
    "fecha": "created_at",
    "created_at": "created_at",
    "tipo": "tipo",
    "cantidad": "cantidad",
    "importe": "cantidad",
    "moneda": "moneda",
    "categoria": "categoria",
    "categoría": "categoria",
    "subcategoria": "subcategoria",
    "subcategoría": "subcategoria",
    "concepto": "concepto",
    "descripcion": "concepto",
    "descripción": "concepto",
    "metodo_pago": "metodo_pago",
    "método_pago": "metodo_pago",
    "metodo de pago": "metodo_pago",
    "cuenta": "cuenta",
    "nota": "nota",
}


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def read_tabular_file(filename: str, content: bytes) -> list[dict[str, Any]]:
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(header or "") for header in rows[0]]
    output: list[dict[str, Any]] = []
    for row in rows[1:]:
        output.append({headers[index]: row[index] if index < len(row) else None for index in range(len(headers))})
    return output


def map_input_row(row: dict[str, Any]) -> tuple[MovementPayload | None, str | None]:
    mapped: dict[str, Any] = {}
    for key, value in row.items():
        field = FIELD_ALIASES.get(normalize_header(key))
        if field:
            mapped[field] = value

    try:
        if "tipo" not in mapped or not mapped["tipo"]:
            mapped["tipo"] = "gasto"
        mapped["tipo"] = str(mapped["tipo"]).strip().lower()
        if mapped["tipo"] not in {"gasto", "ingreso"}:
            return None, "tipo debe ser gasto o ingreso"

        if "cantidad" not in mapped or mapped["cantidad"] in (None, ""):
            return None, "cantidad es obligatoria"
        mapped["cantidad"] = Decimal(str(mapped["cantidad"]).replace(",", "."))

        if mapped.get("created_at"):
            mapped["created_at"] = parse_datetime(mapped["created_at"])

        if not mapped.get("moneda"):
            mapped["moneda"] = "EUR"

        return MovementPayload(**mapped), None
    except (InvalidOperation, ValueError) as exc:
        return None, str(exc)


def parse_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return datetime.fromisoformat(text)


def preview_import(filename: str, content: bytes) -> dict[str, Any]:
    rows = read_tabular_file(filename, content)
    preview_rows = []
    errors = []

    for index, row in enumerate(rows[:50], start=2):
        payload, error = map_input_row(row)
        if error:
            errors.append({"row": index, "error": error})
        elif payload:
            preview_rows.append(payload.model_dump(mode="json"))

    return {
        "filename": filename,
        "total_rows": len(rows),
        "preview_rows": preview_rows,
        "errors": errors,
    }


def commit_import(request: ImportCommitRequest) -> dict[str, Any]:
    imported = 0
    errors = []

    for index, row in enumerate(request.rows, start=1):
        try:
            create_movement(row)
            imported += 1
        except Exception as exc:
            errors.append({"row": index, "error": str(exc)})

    skipped = len(request.rows) - imported
    job_id = execute_returning_id(
        """
        INSERT INTO import_jobs (filename, imported_rows, skipped_rows, errors_json)
        VALUES (%s, %s, %s, %s)
        """,
        (request.filename, imported, skipped, json.dumps(errors, ensure_ascii=False)),
    )

    return {
        "job_id": job_id,
        "imported_rows": imported,
        "skipped_rows": skipped,
        "errors": errors,
    }
